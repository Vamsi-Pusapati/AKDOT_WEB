# scenarioanalysis/scenarioanalysis_service.py

import os
import re
import pandas as pd
from openpyxl import load_workbook

class CSVProcessingError(Exception):
    """
    Custom exception raised when parsing or processing the CSV fails.
    """
    pass

def is_date_string(value):
    """
    Checks if 'value' looks like '"Saturday, March 1, 2025"' (with quotes).
    """
    if not isinstance(value, str):
        return False
    raw = value.strip().strip('"').strip()
    pattern = r"^[A-Za-z]+,\s[A-Za-z]+\s\d{1,2},\s\d{4}$"
    return bool(re.match(pattern, raw))

def parse_day_month_year(date_str):
    """
    E.g. 'Saturday, March 1, 2025' -> (1, 'March', 2025).
    Returns (day, monthName, year).
    """
    raw = date_str.strip().strip('"')
    parts = raw.split(",")  # ["Saturday", " March 1", " 2025"]
    if len(parts) < 3:
        return None, None, None

    middle = parts[1].strip()  # "March 1"
    middle_parts = middle.split()  # ["March", "1"]
    if len(middle_parts) < 2:
        return None, None, None

    month_name = middle_parts[0]
    try:
        day_of_month = int(middle_parts[1])
    except ValueError:
        day_of_month = None

    try:
        year_val = int(parts[2].strip())
    except ValueError:
        year_val = None

    return day_of_month, month_name, year_val

def parse_hour(h):
    """
    Convert '00:00:00' -> 0, etc. Fallback=0 on error.
    """
    if isinstance(h, int):
        return h
    if isinstance(h, float):
        return int(h)
    if isinstance(h, str) and ":" in h:
        # If that fails, we default to 0
        try:
            return int(h.split(":")[0])
        except:
            return 0
    return 0

def determine_sheet_name(df: pd.DataFrame) -> str:
    """
    Look at df['Date'][0], parse out month-year => e.g. "Mar-2025".
    If anything fails or df is empty => "Unknown".
    """
    if df.empty:
        return "Unknown"

    first_date_str = df["Date"].iloc[0]
    day, month_name, year_val = parse_day_month_year(first_date_str)
    if not month_name or not year_val:
        return "Unknown"

    short_map = {
        "January": "Jan", "February": "Feb", "March": "Mar", "April": "Apr",
        "May": "May", "June": "Jun", "July": "Jul", "August": "Aug",
        "September": "Sep", "October": "Oct", "November": "Nov", "December": "Dec"
    }
    short_mon = short_map.get(month_name, month_name[:3])

    return f"{short_mon}-{year_val}"

def load_cars_trucks_from_csv(csv_file):
    """
    Reads CSV in "day+header row, then 24 rows" format => [Date, Hour, Cars_total, Trucks_total].
    Could raise CSVProcessingError if something is obviously malformed.
    """
    try:
        df_raw = pd.read_csv(csv_file, header=None)
    except Exception as e:
        raise CSVProcessingError(f"Failed to read CSV: {str(e)}")

    num_rows = len(df_raw)
    if num_rows == 0:
        raise CSVProcessingError("CSV file is empty.")

    car_cols = ["Mcl","Car","LGV","Bus"]
    truck_cols = ["R2X","R3X","R4+X","A4-X","A5X","A6+X","AT5-X","AT6X","AT7+X","UC","CarLGV+T","Rigid+T"]

    all_blocks = []
    i = 0
    while i < num_rows:
        row_values = df_raw.iloc[i].tolist()
        first_cell = row_values[0]

        if is_date_string(str(first_cell)):
            current_date_str = str(first_cell).strip().strip('"').strip()
            headers = [h for h in row_values[1:] if pd.notna(h) and str(h).strip() != ""]

            if not headers:
                raise CSVProcessingError(f"No headers found on row {i} (suspected date row).")

            day_data_start = i + 1
            day_data_end = day_data_start + 24
            if day_data_end > num_rows:
                # We don't have enough rows for 24 hours
                raise CSVProcessingError(
                    f"CSV data is incomplete after row {i}; expected 24 rows for the day block."
                )

            day_block = df_raw.iloc[day_data_start:day_data_end].copy()
            day_block = day_block.iloc[:, :len(headers)]
            day_block.columns = headers
            day_block["Date"] = current_date_str

            # Convert numeric columns (skip first if it's "Hour")
            if len(headers) > 1:
                numeric_cols = headers[1:]
                for col in numeric_cols:
                    # We'll handle parse errors
                    day_block[col] = pd.to_numeric(day_block[col], errors="coerce").fillna(0)

            all_blocks.append(day_block)
            i = day_data_end
        else:
            i += 1

    if not all_blocks:
        # Means we never found a row that looked like a date+header
        # Possibly the CSV doesn't follow the required format.
        return pd.DataFrame()

    df = pd.concat(all_blocks, ignore_index=True)

    if df.empty:
        return df

    first_col = df.columns[0]
    if first_col not in ["Hour", "Date"]:
        df.rename(columns={first_col: "Hour"}, inplace=True)

    # Convert hour
    df["Hour"] = df["Hour"].apply(parse_hour).astype(int)

    # Summations
    existing_car_cols = [c for c in car_cols if c in df.columns]
    existing_truck_cols = [t for t in truck_cols if t in df.columns]

    df["Cars_total"] = df[existing_car_cols].sum(axis=1)
    df["Trucks_total"] = df[existing_truck_cols].sum(axis=1)

    return df[["Date", "Hour", "Cars_total", "Trucks_total"]]

def pivot_data(df: pd.DataFrame, value_col: str) -> pd.DataFrame:
    """
    Pivot so rows=Hour (0..23), columns=Day(1..31), values=value_col => fill 0.
    """
    def get_day(d):
        day, m, y = parse_day_month_year(d)
        return day

    df = df.copy()
    df["Day"] = df["Date"].apply(get_day)

    pt = df.pivot_table(
        index="Hour",
        columns="Day",
        values=value_col,
        aggfunc="sum",
        fill_value=0
    )

    all_hours = range(24)
    all_days = range(1, 32)
    pt = pt.reindex(index=all_hours, columns=all_days, fill_value=0)
    return pt

def update_excel_sheet(excel_path: str, sheet_name: str, data_frame: pd.DataFrame):
    """
    Remove sheet if exists, then write data_frame with no row/col headers.
    """
    file_exists = os.path.isfile(excel_path)
    if file_exists:
        wb = load_workbook(excel_path)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            wb.remove(ws)
        wb.save(excel_path)
        wb.close()

    mode = "a" if os.path.isfile(excel_path) else "w"
    with pd.ExcelWriter(excel_path, engine="openpyxl", mode=mode) as writer:
        data_frame.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

def process_csv_and_update_excels(csv_path: str) -> str:
    """
    - Parse CSV
    - Determine sheet name => e.g. "Mar-2025"
    - Pivot cars/trucks
    - Overwrite sheets in AKDOTCars.xlsx & AKDOTTrucks.xlsx
    Returns the sheet_name used or raises CSVProcessingError.
    """
    df = load_cars_trucks_from_csv(csv_path)
    if df.empty:
        # We consider "empty final data" an error
        return "NoData"

    sheet_name = determine_sheet_name(df)
    if sheet_name == "Unknown":
        raise CSVProcessingError("Could not determine month/year from the CSV data.")

    cars_pt = pivot_data(df, "Cars_total")
    trucks_pt = pivot_data(df, "Trucks_total")

    folder = os.path.dirname(os.path.abspath(__file__))
    akdot_cars = os.path.join(folder, "AKDOTCars.xlsx")
    akdot_trucks = os.path.join(folder, "AKDOTTrucks.xlsx")

    update_excel_sheet(akdot_cars, sheet_name, cars_pt)
    update_excel_sheet(akdot_trucks, sheet_name, trucks_pt)

    return sheet_name
